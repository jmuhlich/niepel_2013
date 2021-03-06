from __future__ import print_function
from niepel_2013 import *
import openpyxl
import jinja2
import requests
import os
import re


cellline_path = resource_path('SignalingPage', 'CellLinePage')

template_env = jinja2.Environment(
    loader=jinja2.PackageLoader('niepel_2013', 'templates'))
cellline_template = template_env.get_template('cell_line.html')

image_dirs = [
    ('foldchange', 'FoldChangeBoxPlot'),
    ('nodeedge', 'NodeEdgeFigures'),
    ('subtype', 'subtypeBoxPlot'),
    ('topmeasures', 'BasalTopMeasures'),
    ]

image_sizes = {
    'foldchange': [700, 237],
    'nodeedge': [250, 235],
    'subtype': [700, 237],
    'topmeasures': [700, 385],
    }
# Set up dimensions for _large images - start with x2 for all images.
image_sizes_large = dict((name, [width * 2, height * 2])
                         for name, [width, height] in image_sizes.items())
# Override with x3 for nodeedge.
image_sizes_large['nodeedge'] = [dim*3 for dim in image_sizes['nodeedge']]
# Override height for topmeasures, which uses a wholly different image for the
# large size.
image_sizes_large['topmeasures'][1] = 1867

topmeasures_large_suffix = '_allRTKs'

img_path_elements = ('explore', 'cell_line', 'img')
html_path = create_output_path(*img_path_elements[:-1])

print_partial('cell line info')
cellline_info = stash_get('cellline_info')
if not cellline_info:
    filename = os.path.join(cellline_path, 'CellLine_info.xlsx')
    workbook = openpyxl.load_workbook(filename, use_iterators=True)
    sheet = workbook.worksheets[0]
    sheet_iter = sheet.iter_rows()
    sheet_iter.next() # skip header row (iter_rows API for offsets is buggy)
    cellline_info = [[cell.internal_value for cell in row] for row in sheet_iter]
    stash_put('cellline_info', cellline_info)
PASS_nl()

column_names = ('hmsl_id', 'name', 'short_name', 'atcc_id_ignored',
                'vendor', 'class_neve', 'class_heiser', 'class_kao', 'notes',
                'class_consensus', 'growth_medium', 'culture_temperature',
                'culture_atmosphere')

all_data = []
print()
print(' ' * 15 + 'FC NE ST TM DB  status')

for row in cellline_info:

    data = dict(zip(column_names, row))
    print('%-15s' % data['name'], end='')

    plot_filename = data['name'] + '.png'
    all_ok = all([print_status_accessible(cellline_path, d_in, plot_filename)
                  for d_out, d_in in image_dirs])

    if all_ok:
        db_stash_key = 'hmsl_db/%s' % data['hmsl_id']
        db_response = stash_get(db_stash_key)
        if not db_response:
            db_url = 'http://lincs.hms.harvard.edu/db/api/v1/cell/%s/' % data['hmsl_id']
            db_response = requests.get(db_url)
            db_response.raw = None  # hack to make it picklable
            stash_put(db_stash_key, db_response)
        if db_response.ok:
            PASS()
            data['db'] = db_response.json()
            if data['db']['clAlternateID']:
                cosmic_match = re.search(r'COSMIC:\s*(\d+)', data['db']['clAlternateID'])
                if cosmic_match:
                    data['db']['_cosmic_id'] = cosmic_match.group(1)
        else:
            FAIL()
            all_ok = False

    if all_ok:
        print(' OK')
        all_data.append(data)
    else:
        print()

stash_put('cell_lines', all_data)

print()
common = {
          'all_names': [data['name'] for data in all_data],
          'image_sizes': image_sizes,
          'STATIC_URL_2': '../.etc/',
          'DOCROOT': '../../',
         }
breadcrumb_base = [
    {'url': '../../start.html', 'text': 'Start'},
    {'url': None, 'text': 'Cell lines'},
]
for i, data in enumerate(all_data):
    msg = 'rendering page %d/%d %s...' % (i+1, len(all_data), data['name'])
    # FIXME The string padding (37) should be calculated dynamically.
    print_partial('\r' + msg.ljust(37))
    data.update(common)
    data['breadcrumbs'] = breadcrumb_base + [{'url': '', 'text': data['name']}]
    html_filename = data['name'] + '.html'
    render_template(cellline_template, data, html_path, html_filename)
    image_filename = data['name'] + '.png'
    copy_images(image_dirs, image_filename,
                cellline_path, img_path_elements,
                new_sizes=image_sizes, new_format='jpg',
                format_options={'quality': 85, 'optimize': True})
    (if_root, if_ext) = os.path.splitext(image_filename)
    if_root + '_large' + if_ext
    image_dirs_no_topmeasures = [d for d in image_dirs if d[0] != 'topmeasures']
    copy_images(image_dirs_no_topmeasures, image_filename,
                cellline_path, img_path_elements,
                new_sizes=image_sizes_large, new_format='jpg',
                dest_suffix='_large',
                format_options={'quality': 75, 'optimize': True})
    image_dirs_topmeasures = [d for d in image_dirs if d[0] == 'topmeasures']
    if_root, if_ext = os.path.splitext(image_filename)
    image_filename_topmeasures = if_root + topmeasures_large_suffix + if_ext
    copy_images(image_dirs_topmeasures, image_filename_topmeasures,
                cellline_path, img_path_elements,
                new_sizes=image_sizes_large, new_format='jpg',
                dest_suffix='_large',
                format_options={'quality': 75, 'optimize': True})
print_partial("done")
PASS_nl()

print()
subtypes = set(d['class_consensus'] for d in all_data)
for subtype in subtypes:
    image_filename = 'NetMap_%s.png' % subtype
    # We are only copying one image, but we can reuse copy_images with a little
    # creativity in crafting the first arg.
    copy_images([('nodeedge', 'NodeEdgeFigures')], image_filename,
                cellline_path, img_path_elements,
                new_sizes=image_sizes, new_format='jpg',
                format_options={'quality': 85, 'optimize': True})
    copy_images([('nodeedge', 'NodeEdgeFigures')], image_filename,
                cellline_path, img_path_elements,
                new_sizes=image_sizes_large, new_format='jpg',
                dest_suffix='_large',
                format_options={'quality': 85, 'optimize': True})
    print(image_filename, end=' ')
    PASS_nl()
